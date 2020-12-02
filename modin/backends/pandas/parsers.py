# Licensed to Modin Development Team under one or more contributor license agreements.
# See the NOTICE file distributed with this work for additional information regarding
# copyright ownership.  The Modin Development Team licenses this file to you under the
# Apache License, Version 2.0 (the "License"); you may not use this file except in
# compliance with the License.  You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software distributed under
# the License is distributed on an "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF
# ANY KIND, either express or implied. See the License for the specific language
# governing permissions and limitations under the License.

from collections import OrderedDict
from io import BytesIO
import numpy as np
import pandas
from pandas.core.dtypes.cast import find_common_type
from pandas.core.dtypes.concat import union_categoricals
from pandas.io.common import infer_compression
import warnings

from modin.engines.base.io import FileDispatcher
from modin.data_management.utils import split_result_of_axis_func_pandas
from modin.error_message import ErrorMessage


def _split_result_for_readers(axis, num_splits, df):  # pragma: no cover
    """Splits the DataFrame read into smaller DataFrames and handles all edge cases.

    Args:
        axis: Which axis to split over.
        num_splits: The number of splits to create.
        df: The DataFrame after it has been read.

    Returns:
        A list of pandas DataFrames.
    """
    splits = split_result_of_axis_func_pandas(axis, num_splits, df)
    if not isinstance(splits, list):
        splits = [splits]
    return splits


def find_common_type_cat(types):
    if all(isinstance(t, pandas.CategoricalDtype) for t in types):
        if all(t.ordered for t in types):
            return pandas.CategoricalDtype(
                np.sort(np.unique([c for t in types for c in t.categories])[0]),
                ordered=True,
            )
        return union_categoricals(
            [pandas.Categorical([], dtype=t) for t in types],
            sort_categories=all(t.ordered for t in types),
        ).dtype
    else:
        return find_common_type(types)


class PandasParser(object):
    @classmethod
    def get_dtypes(cls, dtypes_ids):
        return (
            pandas.concat(cls.materialize(dtypes_ids), axis=1)
            .apply(lambda row: find_common_type_cat(row.values), axis=1)
            .squeeze(axis=0)
        )

    @classmethod
    def single_worker_read(cls, fname, **kwargs):
        ErrorMessage.default_to_pandas("Parameters provided")
        # Use default args for everything
        pandas_frame = cls.parse(fname, **kwargs)
        if isinstance(pandas_frame, pandas.io.parsers.TextFileReader):
            pd_read = pandas_frame.read
            pandas_frame.read = (
                lambda *args, **kwargs: cls.query_compiler_cls.from_pandas(
                    pd_read(*args, **kwargs), cls.frame_cls
                )
            )
            return pandas_frame
        elif isinstance(pandas_frame, (OrderedDict, dict)):
            return {
                i: cls.query_compiler_cls.from_pandas(frame, cls.frame_cls)
                for i, frame in pandas_frame.items()
            }
        return cls.query_compiler_cls.from_pandas(pandas_frame, cls.frame_cls)

    infer_compression = infer_compression


class PandasCSVParser(PandasParser):
    @staticmethod
    def parse(fname, **kwargs):
        warnings.filterwarnings("ignore")
        num_splits = kwargs.pop("num_splits", None)
        start = kwargs.pop("start", None)
        end = kwargs.pop("end", None)
        index_col = kwargs.get("index_col", None)
        if start is not None and end is not None:
            # pop "compression" from kwargs because bio is uncompressed
            bio = FileDispatcher.file_open(
                fname, "rb", kwargs.pop("compression", "infer")
            )
            if kwargs.get("encoding", None) is not None:
                header = b"" + bio.readline()
            else:
                header = b""
            bio.seek(start)
            to_read = header + bio.read(end - start)
            bio.close()
            pandas_df = pandas.read_csv(BytesIO(to_read), **kwargs)
        else:
            # This only happens when we are reading with only one worker (Default)
            return pandas.read_csv(fname, **kwargs)
        if index_col is not None:
            index = pandas_df.index
        else:
            # The lengths will become the RangeIndex
            index = len(pandas_df)
        return _split_result_for_readers(1, num_splits, pandas_df) + [
            index,
            pandas_df.dtypes,
        ]


class PandasFWFParser(PandasParser):
    @staticmethod
    def parse(fname, **kwargs):
        num_splits = kwargs.pop("num_splits", None)
        start = kwargs.pop("start", None)
        end = kwargs.pop("end", None)
        index_col = kwargs.get("index_col", None)
        if start is not None and end is not None:
            # pop "compression" from kwargs because bio is uncompressed
            bio = FileDispatcher.file_open(
                fname, "rb", kwargs.pop("compression", "infer")
            )
            if kwargs.get("encoding", None) is not None:
                header = b"" + bio.readline()
            else:
                header = b""
            bio.seek(start)
            to_read = header + bio.read(end - start)
            bio.close()
            pandas_df = pandas.read_fwf(BytesIO(to_read), **kwargs)
        else:
            # This only happens when we are reading with only one worker (Default)
            return pandas.read_fwf(fname, **kwargs)
        if index_col is not None:
            index = pandas_df.index
        else:
            # The lengths will become the RangeIndex
            index = len(pandas_df)
        return _split_result_for_readers(1, num_splits, pandas_df) + [
            index,
            pandas_df.dtypes,
        ]


class PandasExcelParser(PandasParser):
    @classmethod
    def get_sheet_data(cls, sheet, convert_float):
        return [
            [cls._convert_cell(cell, convert_float) for cell in row]
            for row in sheet.rows
        ]

    @classmethod
    def _convert_cell(cls, cell, convert_float):
        if cell.is_date:
            return cell.value
        elif cell.data_type == "e":
            return np.nan
        elif cell.data_type == "b":
            return bool(cell.value)
        elif cell.value is None:
            return ""
        elif cell.data_type == "n":
            if convert_float:
                val = int(cell.value)
                if val == cell.value:
                    return val
            else:
                return float(cell.value)

        return cell.value

    @staticmethod
    def parse(fname, **kwargs):
        num_splits = kwargs.pop("num_splits", None)
        start = kwargs.pop("start", None)
        end = kwargs.pop("end", None)
        _skiprows = kwargs.pop("skiprows")
        excel_header = kwargs.get("_header")
        sheet_name = kwargs.get("sheet_name", 0)
        footer = b"</sheetData></worksheet>"

        # Default to pandas case, where we are not splitting or partitioning
        if start is None or end is None:
            return pandas.read_excel(fname, **kwargs)

        from zipfile import ZipFile
        from openpyxl import load_workbook
        from openpyxl.worksheet._reader import WorksheetReader
        from openpyxl.reader.excel import ExcelReader
        from openpyxl.worksheet.worksheet import Worksheet
        from pandas.core.dtypes.common import is_list_like
        from pandas.io.excel._util import (
            _fill_mi_header,
            _maybe_convert_usecols,
        )
        from pandas.io.parsers import TextParser
        import re

        wb = load_workbook(filename=fname, read_only=True)
        # Get shared strings
        ex = ExcelReader(fname, read_only=True)
        ex.read_manifest()
        ex.read_strings()
        # Convert string name 0 to string
        if sheet_name == 0:
            sheet_name = wb.sheetnames[sheet_name]
        # get the worksheet to use with the worksheet reader
        ws = Worksheet(wb)
        # Read the raw data
        with ZipFile(fname) as z:
            with z.open("xl/worksheets/{}.xml".format(sheet_name)) as file:
                file.seek(start)
                bytes_data = file.read(end - start)

        def update_row_nums(match):
            """Update the row numbers to start at 1.

            Note: This is needed because the parser we are using does not scale well if
            the row numbers remain because empty rows are inserted for all "missing"
            rows.

            Parameters
            ----------
            match
                The match from the origin `re.sub` looking for row number tags.

            Returns
            -------
            string
                The updated string with new row numbers.
            """
            b = match.group(0)
            return re.sub(
                b"\d+",  # noqa: W605
                lambda c: str(int(c.group(0).decode("utf-8")) - _skiprows).encode(
                    "utf-8"
                ),
                b,
            )

        bytes_data = re.sub(b'r="[A-Z]*\d+"', update_row_nums, bytes_data)  # noqa: W605
        bytesio = BytesIO(excel_header + bytes_data + footer)
        # Use openpyxl to read/parse sheet data
        reader = WorksheetReader(ws, bytesio, ex.shared_strings, False)
        # Attach cells to worksheet object
        reader.bind_cells()
        data = PandasExcelParser.get_sheet_data(ws, kwargs.pop("convert_float", True))
        usecols = _maybe_convert_usecols(kwargs.pop("usecols", None))
        header = kwargs.pop("header", 0)
        index_col = kwargs.pop("index_col", None)
        # skiprows is handled externally
        skiprows = None

        # Handle header and create MultiIndex for columns if necessary
        if is_list_like(header) and len(header) == 1:
            header = header[0]
        if header is not None and is_list_like(header):
            control_row = [True] * len(data[0])

            for row in header:
                data[row], control_row = _fill_mi_header(data[row], control_row)
        # Handle MultiIndex for row Index if necessary
        if is_list_like(index_col):
            # Forward fill values for MultiIndex index.
            if not is_list_like(header):
                offset = 1 + header
            else:
                offset = 1 + max(header)

            # Check if dataset is empty
            if offset < len(data):
                for col in index_col:
                    last = data[offset][col]
                    for row in range(offset + 1, len(data)):
                        if data[row][col] == "" or data[row][col] is None:
                            data[row][col] = last
                        else:
                            last = data[row][col]

        parser = TextParser(
            data,
            header=header,
            index_col=index_col,
            has_index_names=is_list_like(header) and len(header) > 1,
            skiprows=skiprows,
            usecols=usecols,
            **kwargs
        )
        # In excel if you create a row with only a border (no values), this parser will
        # interpret that as a row of NaN values. Pandas discards these values, so we
        # also must discard these values.
        pandas_df = parser.read().dropna(how="all")
        # Since we know the number of rows that occur before this partition, we can
        # correctly assign the index in cases of RangeIndex. If it is not a RangeIndex,
        # the index is already correct because it came from the data.
        if isinstance(pandas_df.index, pandas.RangeIndex):
            pandas_df.index = pandas.RangeIndex(
                start=_skiprows, stop=len(pandas_df.index) + _skiprows
            )
        # We return the length if it is a RangeIndex (common case) to reduce
        # serialization cost.
        if index_col is not None:
            index = pandas_df.index
        else:
            # The lengths will become the RangeIndex
            index = len(pandas_df)
        return _split_result_for_readers(1, num_splits, pandas_df) + [
            index,
            pandas_df.dtypes,
        ]


class PandasJSONParser(PandasParser):
    @staticmethod
    def parse(fname, **kwargs):
        num_splits = kwargs.pop("num_splits", None)
        start = kwargs.pop("start", None)
        end = kwargs.pop("end", None)
        if start is not None and end is not None:
            # pop "compression" from kwargs because bio is uncompressed
            bio = FileDispatcher.file_open(
                fname, "rb", kwargs.pop("compression", "infer")
            )
            bio.seek(start)
            to_read = b"" + bio.read(end - start)
            bio.close()
            columns = kwargs.pop("columns")
            pandas_df = pandas.read_json(BytesIO(to_read), **kwargs)
        else:
            # This only happens when we are reading with only one worker (Default)
            return pandas.read_json(fname, **kwargs)
        if not pandas_df.columns.equals(columns):
            raise NotImplementedError("Columns must be the same across all rows.")
        partition_columns = pandas_df.columns
        return _split_result_for_readers(1, num_splits, pandas_df) + [
            len(pandas_df),
            pandas_df.dtypes,
            partition_columns,
        ]


class PandasParquetParser(PandasParser):
    @staticmethod
    def parse(fname, **kwargs):
        num_splits = kwargs.pop("num_splits", None)
        columns = kwargs.get("columns", None)
        if fname.startswith("s3://"):
            from botocore.exceptions import NoCredentialsError
            import s3fs

            try:
                fs = s3fs.S3FileSystem()
                fname = fs.open(fname)
            except NoCredentialsError:
                fs = s3fs.S3FileSystem(anon=True)
                fname = fs.open(fname)

        if num_splits is None:
            return pandas.read_parquet(fname, **kwargs)
        kwargs["use_pandas_metadata"] = True
        df = pandas.read_parquet(fname, **kwargs)
        if isinstance(df.index, pandas.RangeIndex):
            idx = len(df.index)
        else:
            idx = df.index
        columns = [c for c in columns if c not in df.index.names and c in df.columns]
        if columns is not None:
            df = df[columns]
        # Append the length of the index here to build it externally
        return _split_result_for_readers(0, num_splits, df) + [idx, df.dtypes]


class PandasHDFParser(PandasParser):  # pragma: no cover
    @staticmethod
    def parse(fname, **kwargs):
        kwargs["key"] = kwargs.pop("_key", None)
        num_splits = kwargs.pop("num_splits", None)
        if num_splits is None:
            return pandas.read_hdf(fname, **kwargs)
        df = pandas.read_hdf(fname, **kwargs)
        # Append the length of the index here to build it externally
        return _split_result_for_readers(0, num_splits, df) + [len(df.index), df.dtypes]


class PandasFeatherParser(PandasParser):
    @staticmethod
    def parse(fname, **kwargs):
        from pyarrow import feather

        num_splits = kwargs.pop("num_splits", None)
        if num_splits is None:
            return pandas.read_feather(fname, **kwargs)
        df = feather.read_feather(fname, **kwargs)
        # Append the length of the index here to build it externally
        return _split_result_for_readers(0, num_splits, df) + [len(df.index), df.dtypes]


class PandasSQLParser(PandasParser):
    @staticmethod
    def parse(sql, con, index_col, **kwargs):
        num_splits = kwargs.pop("num_splits", None)
        if num_splits is None:
            return pandas.read_sql(sql, con, index_col=index_col, **kwargs)
        df = pandas.read_sql(sql, con, index_col=index_col, **kwargs)
        if index_col is None:
            index = len(df)
        else:
            index = df.index
        return _split_result_for_readers(1, num_splits, df) + [index, df.dtypes]
